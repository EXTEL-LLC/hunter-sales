"""
export_for_web.py
─────────────────────────────────────────────────────────────────────────────
Читає EMOS_Pricing_MVP.xlsx та генерує файли для веб-розгортання:

  docs/skus.json      — всі SKU (ціни, категорії, канали)
  docs/version.json   — мітка версії для автооновлення у менеджерів
  docs/index.html     — копія EMOS_Order_Form.html

Папка docs/ публікується через GitHub Pages автоматично при git push.
Форма менеджерів оновить ціни впродовж 5 хвилин після push.

Запуск: python3 export_for_web.py
"""

import re, json, math, shutil, os, datetime
from openpyxl import load_workbook

XLSX      = '/Users/viktorprokopenko/Desktop/CLAUDE/EMOS_Pricing_MVP.xlsx'
HTML      = '/Users/viktorprokopenko/Desktop/CLAUDE/EMOS_Order_Form.html'
WEB       = '/Users/viktorprokopenko/Desktop/CLAUDE/docs'
AUTH_CFG  = '/Users/viktorprokopenko/Desktop/CLAUDE/auth_config.json'

os.makedirs(WEB, exist_ok=True)

wb      = load_workbook(XLSX, data_only=True)
ws_st   = wb['Settings']
ws_sk   = wb['SKU_Base']

# ── Settings ──────────────────────────────────────────────────────────────────
exchange_rate = ws_st.cell(2, 2).value
b3  = ws_st.cell(3, 2).value
b11 = ws_st.cell(11, 2).value or 0
b12 = ws_st.cell(12, 2).value or 0
b13 = ws_st.cell(13, 2).value or 0
b14 = ws_st.cell(14, 2).value or 0
b15 = ws_st.cell(15, 2).value or 0
res = b11 + b12 + b13 + b14 + b15 * b3 / 365

d6_sub = ws_st.cell(68, 9).value or 0.25   # D6 is now col I (9) after D4NI insert

ROLES       = ['Трафіковий', 'Основний', 'Маржинальний', 'Сезонний', 'Стратегічний']
MARGIN_ROWS = [78, 80, 82, 84, 86]
CHANNELS    = ['D1', 'D2', 'D3', 'D4 Нац', 'D4 Buy', 'D4 Drop', 'D5', 'D6']

MARGINS = {}
for role, row in zip(ROLES, MARGIN_ROWS):
    MARGINS[role] = [ws_st.cell(row, c).value or 0 for c in range(2, 10)]

# Average D4 НІ markup (fallback for SKUs not in markup file)
NI_AVG_MARKUP = 1.3445

ROLE_NT = {
    'Трафіковий':   0.16,
    'Основний':     0.19,
    'Маржинальний': 0.24,
    'Сезонний':     0.23,
    'Стратегічний': 0.28,
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def psych(p):
    if not p or p <= 0: return 0
    p = round(p)
    last = p % 10
    base = (p // 10) * 10
    if last == 0:          return base - 1
    elif 1 <= last <= 5:   return base + 6
    else:                  return base + 9

def ddp_round(val):
    if not val or val <= 0: return 0
    return round(round(val / 6, 2) * 6, 2)

# ── SKU_Base ──────────────────────────────────────────────────────────────────
skus = []
cats_set = []

for r in range(2, ws_sk.max_row + 1):
    code = ws_sk.cell(r, 2).value
    if not code: continue

    name       = ws_sk.cell(r, 3).value or ''
    category   = ws_sk.cell(r, 4).value or ''
    unit_raw   = ws_sk.cell(r, 5).value
    invoice    = ws_sk.cell(r, 6).value or 0
    imp_coef   = ws_sk.cell(r, 8).value or 1
    rrp_manual = ws_sk.cell(r, 19).value
    role       = ws_sk.cell(r, 21).value or 'Основний'
    if role not in MARGINS: role = 'Основний'
    ni_markup  = ws_sk.cell(r, 48).value or None

    try:
        unit_mult = float(unit_raw) if unit_raw and float(unit_raw) > 1 else 1
    except (TypeError, ValueError):
        unit_mult = 1

    unit_str = str(unit_raw) if unit_raw else '1'
    cost = invoice * exchange_rate * imp_coef * unit_mult if invoice else 0

    if rrp_manual and isinstance(rrp_manual, (int, float)) and rrp_manual > 0:
        rrp_eff = psych(float(rrp_manual))
    else:
        nt    = ROLE_NT.get(role, 0.19)
        cli1  = MARGINS[role][0]
        denom = (1 - nt) * (1 - res) * (1 - cli1)
        rrp_eff = psych(round(cost / denom, 0)) if denom > 0 and cost > 0 else 0

    margins = MARGINS[role]
    ddp_all = []
    for ch_i, ch_name in enumerate(CHANNELS):
        m = margins[ch_i]
        if ch_name == 'D5' or m == 0:
            ddp_all.append(0)
        elif ch_name == 'D6':
            ddp_all.append(ddp_round(rrp_eff * (1 - m) * (1 - d6_sub)))
        elif ch_name == 'D4 Нац':
            mu = ni_markup if ni_markup else NI_AVG_MARKUP
            ddp_all.append(ddp_round(rrp_eff / mu) if rrp_eff > 0 else 0)
        else:
            ddp_all.append(ddp_round(rrp_eff * (1 - m)))

    skus.append([str(code), name, unit_str, round(cost, 2), ddp_all, rrp_eff, category])

    if category and category not in cats_set:
        cats_set.append(category)

print(f'✓ SKU: {len(skus)}  Категорій: {len(cats_set)}')

# ── Commercial_Policy ─────────────────────────────────────────────────────────
policy = []
ws_cp = wb['Commercial_Policy'] if 'Commercial_Policy' in wb.sheetnames else None
if ws_cp:
    last_ch = None
    for r in range(4, min(28, ws_cp.max_row + 1)):
        ch_raw = ws_cp.cell(r, 1).value
        ct_raw = ws_cp.cell(r, 2).value
        if not ct_raw or str(ct_raw).strip() == '': continue
        if ch_raw:
            first_line = str(ch_raw).split('\n')[0].strip()
            if first_line and not first_line.startswith('='): last_ch = first_line
        if not last_ch: continue

        def _num(v): return v if isinstance(v, (int, float)) else 0
        def _str(v): return str(v) if v else '—'

        disc_num  = _num(ws_cp.cell(r, 8).value)
        welc_num  = _num(ws_cp.cell(r, 9).value)
        retro_raw = ws_cp.cell(r, 15).value
        retro_num = _num(retro_raw)
        retro_str = (str(retro_raw) if isinstance(retro_raw, str) and '–' in str(retro_raw)
                     else (f'{round(retro_num*100)}%' if retro_num > 0 else '—'))

        pay_raw = ws_cp.cell(r, 14).value
        policy.append({
            'channel': last_ch,
            'ct':      str(ct_raw),
            'stage':   _str(ws_cp.cell(r, 3).value),
            'scenario':_str(ws_cp.cell(r, 4).value),
            'goal':    _str(ws_cp.cell(r, 5).value),
            'vol_min': _num(ws_cp.cell(r, 6).value),
            'vol_max': _num(ws_cp.cell(r, 7).value) or 999999999,
            'disc':    f'{round(disc_num*100)}%' if disc_num > 0 else '—',
            'welcome': f'{round(welc_num*100)}%' if welc_num > 0 else '—',
            'welcome_orders': _num(ws_cp.cell(r, 10).value),
            'welcome_days':   _num(ws_cp.cell(r, 11).value),
            'delivery': _str(ws_cp.cell(r, 12).value),
            'moq':      _num(ws_cp.cell(r, 13).value),
            'pay_days': pay_raw if isinstance(pay_raw, (int, float)) else None,
            'retro':    retro_str,
            'kpi_cond': _str(ws_cp.cell(r, 16).value),
            'kpi_period': _str(ws_cp.cell(r, 17).value),
            'marketing': _str(ws_cp.cell(r, 18).value),
            'approval':  _str(ws_cp.cell(r, 19).value),
            'rule_code': _str(ws_cp.cell(r, 20).value),
            'note':      str(ws_cp.cell(r, 21).value or ''),
            'max_disc':  round(disc_num * 100),
            'max_welc':  round(welc_num * 100),
            'max_retro': round(retro_num * 100),
        })
    print(f'✓ Policy: {len(policy)} рядків')
else:
    print('⚠ Лист Commercial_Policy не знайдено — policy буде порожнім')

# ── version.json ──────────────────────────────────────────────────────────────
now = datetime.datetime.now()
version_str = now.strftime('%Y%m%d-%H%M')
version_data = {'v': version_str, 'ts': int(now.timestamp())}

with open(f'{WEB}/version.json', 'w', encoding='utf-8') as f:
    json.dump(version_data, f)
print(f'✓ version.json: {version_str}')

# ── skus.json ─────────────────────────────────────────────────────────────────
# ── auth_config.json ──────────────────────────────────────────────────────────
auth_data = {}
if os.path.exists(AUTH_CFG):
    with open(AUTH_CFG, 'r', encoding='utf-8') as f:
        auth_data = json.load(f)
    print(f'✓ auth_config.json: {len(auth_data.get("managers", []))} менеджерів')
else:
    print('⚠ auth_config.json не знайдено — auth не буде в skus.json')

skus_data = {
    'version': version_str,
    'ch_idx': {'D1':0,'D2':1,'D3':2,'D4 Нац':3,'D4 Buy':4,'D4 Drop':5,'D5':6,'D6':7},
    'cats': cats_set,
    'skus': skus,
    'policy': policy,
    'auth': auth_data,
}
with open(f'{WEB}/skus.json', 'w', encoding='utf-8') as f:
    json.dump(skus_data, f, ensure_ascii=False, separators=(',', ':'))
size_kb = os.path.getsize(f'{WEB}/skus.json') // 1024
print(f'✓ skus.json: {len(skus)} SKU, {size_kb} КБ')

# ── index.html — copy ─────────────────────────────────────────────────────────
shutil.copy2(HTML, f'{WEB}/index.html')
print(f'✓ index.html скопійовано')

print(f'\n{"─"*60}')
print(f'  Папка для розгортання: {WEB}')
print(f'  GitHub Pages: git add docs/ && git commit -m "update prices" && git push')
print(f'  або:          bash git_deploy.sh')
print(f'{"─"*60}')
print(f'  Менеджери отримають оновлення автоматично (~5 хв)')
