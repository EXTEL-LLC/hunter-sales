"""
process_feed.py
───────────────────────────────────────────────────────────────────────────────
Обробляє Excel-файли «Асортимент» та «Emos Прайс» і генерує XML-фід
для Telemart.ua у файл web/feed.xml.

Запуск:
    # Перший раз (обидва файли):
    python3 process_feed.py

    # Тільки «Асортимент» (ціни беруться з кешу):
    python3 process_feed.py

Скрипт сам знаходить файли на Робочому столі за маскою імені.
Або передайте шляхи явно:
    python3 process_feed.py --assortment /path/to/file.xlsx --price /path/to/price.xlsx
"""

import os, sys, json, glob, argparse, datetime
from pathlib import Path
from openpyxl import load_workbook

# ── Шляхи ──────────────────────────────────────────────────────────────────────
BASE        = Path(__file__).parent
DESKTOP     = Path.home() / 'Desktop'
CACHE_FILE  = BASE / 'data' / 'price_cache.json'
OUT_XML     = BASE / 'web' / 'feed.xml'
os.makedirs(BASE / 'data', exist_ok=True)
os.makedirs(BASE / 'web',  exist_ok=True)

# ── Маски пошуку файлів ────────────────────────────────────────────────────────
ASSORTMENT_PATTERNS = ['Асортимент*.xlsx', 'Асортимент.xlsx', 'assortment*.xlsx']
PRICE_PATTERNS      = ['EMOS_Прайс*.xlsx', 'Emos_Прайс*.xlsx', 'EMOS_Price*.xlsx',
                        'EMOS Прайс*.xlsx', 'Emos Прайс*.xlsx']

# ── Колонки «Асортимент» (1-based) ────────────────────────────────────────────
KEEP_COLS = [1, 6, 8, 9, 10, 11, 12, 14, 16, 20, 38]

RENAME_MAP = {
    6:  'Назва',
    9:  'Категорія',
    10: 'РРЦ',
    11: 'Ціна дроп',
    14: 'Наявність',
    38: 'Опис',
}

# Фінальні назви колонок після обробки (в тому ж порядку що KEEP_COLS)
COL_NAMES_FINAL = []  # заповнюється при читанні

# ── XML-маппінг ────────────────────────────────────────────────────────────────
XML_MAP = {
    'Артикул':    'article',
    'Назва':      'name',
    'Бренд':      'brand',
    'Категорія':  'category',
    'РРЦ':        'price_rrc',
    'Ціна дроп':  'price',
    'Валюта':     'currencyId',
    'Наявність':  'available',
    'Фото':       'pictures',  # спеціальна обробка
    'Ссылка':     'url',       # колонка 20 = Ссылка
    'Опис':       'description',
}
HTML_COLS  = {'Назва', 'Опис'}
PHOTO_SEP  = ';'
AVAIL_COL  = 'Наявність'
PHOTO_COL  = 'Фото'
CAT_COL    = 'Категорія'
ART_COL    = 'Артикул'

# ── Утиліти ───────────────────────────────────────────────────────────────────
def find_file(patterns, label):
    """Шукає файл за масками на Робочому столі."""
    for p in patterns:
        found = sorted(glob.glob(str(DESKTOP / p)))
        if found:
            print(f'  Знайдено {label}: {Path(found[-1]).name}')
            return found[-1]
    return None

def esc_xml(s):
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

def esc_attr(s):
    return str(s).replace('"','&quot;').replace('<','&lt;').replace('>','&gt;')

def avail_val(raw, fmt='bool'):
    v = str(raw or '').lower()
    ok = 'наявн' in v or v in ('1','true','є в наявності')
    if fmt == 'bool': return 'true' if ok else 'false'
    if fmt == 'int':  return '1'    if ok else '0'
    return str(raw or '')

# ── Крок 1: завантажити кеш цін ───────────────────────────────────────────────
def load_price_cache():
    if CACHE_FILE.exists():
        with open(CACHE_FILE, encoding='utf-8') as f:
            data = json.load(f)
        print(f'  Кеш цін: {len(data)} артикулів (з {CACHE_FILE.name})')
        return data
    print('  Кеш цін: відсутній')
    return {}

def save_price_cache(price_map):
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(price_map, f, ensure_ascii=False, indent=2)
    print(f'  Збережено кеш: {len(price_map)} артикулів → {CACHE_FILE.name}')

# ── Крок 2: зчитати Emos Прайс ────────────────────────────────────────────────
def read_price_file(path):
    """Повертає dict {SKU: DDP_price}. Заголовок прайсу на рядку 7."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    price_map = {}
    # Знаходимо рядок з "DDP" у заголовку
    header_row = 7
    for r in range(1, 15):
        row_vals = [ws.cell(r, c).value for c in range(1, 7)]
        if any('DDP' in str(v or '') for v in row_vals):
            header_row = r
            break
    # Знаходимо індекси колонок SKU та DDP
    headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, 7)]
    sku_col = next((i+1 for i, h in enumerate(headers) if 'SKU' in h.upper() or 'КОД' in h.upper()), 2)
    ddp_col = next((i+1 for i, h in enumerate(headers) if 'DDP' in h.upper()), 5)
    for r in range(header_row + 1, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        ddp = ws.cell(r, ddp_col).value
        if sku and ddp and str(sku).strip():
            try:
                price_map[str(sku).strip()] = float(ddp)
            except (ValueError, TypeError):
                pass
    return price_map

# ── Крок 3: зчитати та обробити Асортимент ────────────────────────────────────
def read_assortment(path, price_cache):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # Оригінальні заголовки
    orig_headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    # Фінальні назви для потрібних колонок
    final_headers = []
    for col_idx in KEEP_COLS:
        orig_name = orig_headers[col_idx - 1] if col_idx <= len(orig_headers) else f'Col{col_idx}'
        final_headers.append(RENAME_MAP.get(col_idx, orig_name))
    art_local = final_headers[KEEP_COLS.index(1)]  # = 'Артикул'
    drop_local = RENAME_MAP[11]  # = 'Ціна дроп'
    # Читаємо рядки
    products = []
    matched = not_found = 0
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, col_name in zip(KEEP_COLS, final_headers):
            row_data[col_name] = ws.cell(r, col_idx).value
        art = str(row_data.get(art_local) or '').strip()
        if not art:
            continue
        # Оновлення ціни дроп з кешу
        if art in price_cache:
            row_data[drop_local] = price_cache[art]
            matched += 1
        else:
            not_found += 1
        products.append(row_data)
    print(f'  Асортимент: {len(products)} товарів, ціни DDP: {matched} знайдено / {not_found} не знайдено')
    return products, final_headers

# ── Крок 4: побудова категорій ────────────────────────────────────────────────
def build_category_map(products):
    seen = list(dict.fromkeys(
        str(p.get(CAT_COL) or '').strip() for p in products
        if p.get(CAT_COL)
    ))
    next_id, parent_ids, id_map, lines = 1, {}, {}, []
    # Батьківські
    for n in seen:
        if '/' in n:
            parent = n.split('/')[0].strip()
            if parent not in parent_ids:
                parent_ids[parent] = next_id
                lines.append(f'    <category id="{next_id}">{esc_xml(parent)}</category>')
                next_id += 1
    # Дочірні + без батьків
    for full in seen:
        if '/' in full:
            parent = full.split('/')[0].strip()
            child  = full.split('/', 1)[1].strip()
            pid    = parent_ids.get(parent, '')
            id_map[full] = next_id
            pid_attr = f' parentId="{pid}"' if pid else ''
            lines.append(f'    <category id="{next_id}"{pid_attr}>{esc_xml(child)}</category>')
        else:
            if full in parent_ids:
                id_map[full] = parent_ids[full]
            else:
                id_map[full] = next_id
                lines.append(f'    <category id="{next_id}">{esc_xml(full)}</category>')
        next_id += 1
    return id_map, lines

# ── Крок 5: генерація XML ─────────────────────────────────────────────────────
def build_xml(products, headers, cat_id_map, cat_lines):
    date_str = datetime.date.today().isoformat()
    lines = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append(f'<price date="{date_str}">')
    lines.append('  <categories>')
    lines.extend(cat_lines)
    lines.append('  </categories>')
    lines.append('  <offers>')
    for p in products:
        art      = str(p.get(ART_COL) or '').strip()
        cat_full = str(p.get(CAT_COL) or '').strip()
        cat_id   = cat_id_map.get(cat_full, '')
        avail    = avail_val(p.get(AVAIL_COL))
        grp_attr = f' group_id="{cat_id}"' if cat_id else ''
        lines.append(f'    <offer id="{esc_attr(art)}" available="{avail}"{grp_attr}>')
        for col_name in headers:
            xml_tag = XML_MAP.get(col_name)
            if not xml_tag:
                continue
            val = str(p.get(col_name) or '').strip()
            if col_name == CAT_COL:
                if cat_id:
                    lines.append(f'      <categoryId>{cat_id}</categoryId>')
            elif col_name == PHOTO_COL:
                pics = [u.strip() for u in val.replace('\n', '').split(PHOTO_SEP) if u.strip()]
                if pics:
                    lines.append('      <pictures>')
                    for pic in pics:
                        lines.append(f'        <picture>{esc_xml(pic)}</picture>')
                    lines.append('      </pictures>')
            elif col_name in HTML_COLS:
                if val:
                    inner = val.replace(']]>', ']]]]><![CDATA[>')
                    lines.append(f'      <{xml_tag}><![CDATA[{inner}]]></{xml_tag}>')
            else:
                if val:
                    lines.append(f'      <{xml_tag}>{esc_xml(val)}</{xml_tag}>')
        lines.append('    </offer>')
    lines.append('  </offers>')
    lines.append('</price>')
    return '\n'.join(lines)

# ── Головна функція ───────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Генерація XML-фіду Telemart')
    parser.add_argument('--assortment', help='Шлях до файлу Асортимент.xlsx')
    parser.add_argument('--price',      help='Шлях до файлу Emos Прайс.xlsx (необов\'язково)')
    args = parser.parse_args()

    print('─' * 60)
    print('  Генерація XML-фіду Telemart')
    print('─' * 60)

    # Знаходимо файли
    assortment_path = args.assortment or find_file(ASSORTMENT_PATTERNS, '«Асортимент»')
    price_path      = args.price      or find_file(PRICE_PATTERNS,      '«Emos Прайс»')

    if not assortment_path:
        print('\n❌ Файл «Асортимент» не знайдено!')
        print('   Покладіть файл на Робочий стіл або вкажіть шлях:')
        print('   python3 process_feed.py --assortment /шлях/до/Асортимент.xlsx')
        sys.exit(1)

    # Завантажуємо кеш цін
    print('\n[1/5] Завантаження кешу цін...')
    price_cache = load_price_cache()

    # Оновлення кешу якщо є прайс-файл
    if price_path:
        print('\n[2/5] Читання цін з «Emos Прайс»...')
        new_prices = read_price_file(price_path)
        print(f'  Знайдено {len(new_prices)} артикулів з DDP-ціною')
        price_cache.update(new_prices)
        save_price_cache(price_cache)
    else:
        print('\n[2/5] «Emos Прайс» не надано — використовуємо кеш')

    # Обробка Асортименту
    print('\n[3/5] Обробка файлу «Асортимент»...')
    products, headers = read_assortment(assortment_path, price_cache)

    # Побудова категорій
    print('\n[4/5] Формування категорій...')
    cat_id_map, cat_lines = build_category_map(products)
    print(f'  Унікальних категорій: {len(cat_id_map)}')
    for ln in cat_lines:
        print(f'  {ln.strip()}')

    # Генерація XML
    print('\n[5/5] Генерація XML...')
    xml_content = build_xml(products, headers, cat_id_map, cat_lines)
    with open(OUT_XML, 'w', encoding='utf-8') as f:
        f.write(xml_content)
    size_kb = OUT_XML.stat().st_size // 1024
    print(f'  ✅ Збережено: {OUT_XML}  ({size_kb} КБ, {len(products)} товарів)')

    print('\n─' * 60)
    print(f'  XML готовий: web/feed.xml')
    print(f'  Для публікації запустіть: bash netlify_deploy.sh')
    print('─' * 60)

if __name__ == '__main__':
    main()
